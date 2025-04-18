from flask import Flask, render_template, request ,send_file
import pandas as pd
import asyncio, os, openpyxl
from twilio.rest import Client
import pymysql,math,datetime
# Twilio account credentials
account_sid = os.getenv("TWILIO_ACCOUNT_SID")
auth_token = os.getenv("TWILIO_AUTH_TOKEN")
twilio_client = Client(account_sid, auth_token)
# Function to read Excel file and convert to array
def read_excel_to_array(file_path):
    df = pd.read_excel(file_path)
    return df.values.tolist()

def header_read(file_path):
    df = pd.read_excel(file_path)
    return df.columns

def columns_read():
    wb = openpyxl.load_workbook('Marks1.xlsx')
    ws = wb.active
    return len(list(ws.iter_cols(values_only=True)))
import openpyxl
from openpyxl.cell import MergedCell

def after_process_ese(file_path):
    # Load the workbook and select the active sheet
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Step 1: Unmerge all merged cells (if you want to modify merged cells)
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        ws.unmerge_cells(str(merged_range))

    # Step 2: Process cells after unmerging
    for row in ws.iter_rows():
        for cell in row:
            # Check if the cell is not a MergedCell and set its value to None
            if not isinstance(cell, MergedCell):
                cell.value = None

    # Step 3: Save and close the workbook
    wb.save(file_path)
    wb.close()
def after_process():
    wb = openpyxl.load_workbook('Marks1.xlsx') 
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
            wb.save('Marks1.xlsx')
    return None
async def login_main(login,email,password):
    hod_email=os.getenv("HOD_EMAIL")
    hod_pwd=os.getenv("HOD_PWD")
    staff_email=os.getenv("STAFF_EMAIL")
    staff_pwd=os.getenv("STAFF_PWD")
    if str(login)=="HOD" and str(email)==hod_email and str(password)==hod_pwd:
        stat="hod"
        return True
    elif str(login)=="Staff" and str(email)==staff_email and str(password)==staff_pwd:
        stat="staff"
        return False
    else:
        stat='none'
        return stat
def sendadmin1_msg(message,ph_no):
    try:
        message = twilio_client.messages.create(
            from_='+15392212587',
            to=f"{ph_no}",
            body=message
        )
        print(f"Message sent to {ph_no} regarding arrears.")
    except Exception as e:
        print(f"Failed to send message to {ph_no}: {str(e)}")
def sendadmin2_msg(message,ph_no):
    try:
        message = twilio_client.messages.create(
            from_='+15392212587',
            to=f"{ph_no}",
            body=message
        )
        print(f"Message sent to {ph_no} regarding arrears.")
    except Exception as e:
        print(f"Failed to send message to {ph_no}: {str(e)}")
async def send_sms_message(name,count,sem,exam,year,ph_no, message, cursor, cnx):
    try:
        message = twilio_client.messages.create(
            from_='+15392212587',
            to=f"{ph_no}",
            body=message
        )
        query="use status"
        cursor.execute(query)
        status="DONE"
        query1="insert into status_data(name,arrear_count,sem,exam,year,status) values (%s,%s,%s,%s,%s,%s)"
        values=[name,count,sem,exam,year,status]
        cursor.execute(query1,values)
        cnx.commit()
        query2="use status_rec"
        cursor.execute(query2)
        status="DONE"
        now = datetime.datetime.now()
        query3="insert into status_data(name,arrear_count,sem,exam,year,status,date) values (%s,%s,%s,%s,%s,%s,%s)"
        values=[name,count,sem,exam,year,status,now]
        cursor.execute(query3,values)
        cnx.commit()
        print(f"Message sent to {ph_no} regarding arrears.")
    except Exception as e:
        query="use status"
        cursor.execute(query)
        status="PENDING"
        query1="insert into status_data(name,arrear_count,sem,exam,year,status) values (%s,%s,%s,%s,%s,%s)"
        values=[name,count,sem,exam,year,status]
        cursor.execute(query1,values)
        cnx.commit()
        query2="use status_rec"
        cursor.execute(query2)
        status="PENDING"
        now = datetime.datetime.now()
        query3="insert into status_data(name,arrear_count,sem,exam,year,status,date) values (%s,%s,%s,%s,%s,%s,%s)"
        values=[name,count,sem,exam,year,status,now]
        cursor.execute(query3,values)
        cnx.commit()
        print(f"Failed to send message to {ph_no}: {str(e)}")
async def send_sms_message1(name,count,exam,year,ph_no, message, cursor, cnx):
    try:
        message = twilio_client.messages.create(
            from_='+15392212587',
            to=f"{ph_no}",
            body=message
        )
        query="use overall_rec"
        cursor.execute(query)
        status="DONE"
        query1="insert into overall_data(name,arrear_count,exam,year,Status) values (%s,%s,%s,%s,%s)"
        values=[name,count,exam,year,status]
        cursor.execute(query1,values)
        cnx.commit()
        query2="use overall_rec"
        cursor.execute(query2)
        status="DONE"
        now = datetime.datetime.now()
        query3="insert into overall_data(name,arrear_count,exam,year,Status,DATE) values (%s,%s,%s,%s,%s,%s)"
        values=[name,count,exam,year,status,now]
        cursor.execute(query3,values)
        cnx.commit()
        print(f"Message sent to {ph_no} regarding arrears.")
    except Exception as e:
        query="use overall"
        cursor.execute(query)
        status="PENDING"
        query1="insert into overall_data(name,arrear_count,exam,year,Status) values (%s,%s,%s,%s,%s)"
        values=[name,count,exam,year,status]
        cursor.execute(query1,values)
        cnx.commit()
        query2="use overall_rec"
        cursor.execute(query2)
        status="PENDING"
        now = datetime.datetime.now()
        query3="insert into overall_data(name,arrear_count,exam,year,Status,DATE) values (%s,%s,%s,%s,%s,%s)"
        values=[name,count,exam,year,status,now]
        cursor.execute(query3,values)
        cnx.commit()
        print(f"Failed to send message to {ph_no}: {str(e)}")
def process_hod_data_overall(year,exam,arrear,cnx,cursor):
    data = None  # Initialize `data` to avoid UnboundLocalError
        # Mapping arrear type to database name
    if arrear == 'five_and_above':
        cursor.execute("USE 5_arrear_data")
        query = "SELECT name, arrear_count,year,exam FROM 5_arrear WHERE year = %s  AND exam = %s"
        cursor.execute(query, (year, exam))
        data = cursor.fetchall()
    else:
        print("Invalid arrear type")
    return data
def process_hod_data(year, sem, exam, arrear,cnx,cursor):
    data = None  # Initialize `data` to avoid UnboundLocalError
    if arrear == '3_and_above_arrear':
        cursor.execute("USE 3_arrear_data")
        query = "SELECT name, arrear_count,year,sem,exam FROM 3_arrear WHERE year = %s AND sem = %s AND exam = %s"
        cursor.execute(query, (year, sem, exam))
        data = cursor.fetchall()
    elif arrear == 'two_arrear':
        cursor.execute("USE 2_arrear_data")
        query = "SELECT name, arrear_count,year,sem,exam FROM 2_arrear WHERE year = %s AND sem = %s AND exam = %s"
        cursor.execute(query, (year, sem, exam))
        data = cursor.fetchall()
    elif arrear == 'one_arrear':
        cursor.execute("USE 1_arrear_data")
        query = "SELECT name, arrear_count,year,sem,exam FROM 1_arrear WHERE year = %s AND sem = %s AND exam = %s"
        cursor.execute(query, (year, sem, exam))
        data = cursor.fetchall()
    elif arrear == 'nil_arrear':
        cursor.execute("USE nil_arrear_data")
        query = "SELECT name, arrear_count,year,sem,exam FROM nil_arrear WHERE year = %s AND sem = %s AND exam = %s"
        cursor.execute(query, (year, sem, exam))
        data = cursor.fetchall()
    else:
        print("Invalid arrear type")
    return data
def clear_data(arrear,year,exam,sem):
    # Establish a connection to the MySQL database
    db_user = os.getenv("DB_USER")
    db_password = os.getenv("DB_PASSWORD")
    db_host = os.getenv("DB_HOST")
    cnx = pymysql.connect(
    cursorclass=pymysql.cursors.DictCursor,
    host=db_host,
    password=db_password,
    port=15274,
    user=db_user,)
    cursor = cnx.cursor()
    try:
        # Mapping arrear type to database name
        if arrear == '3_and_above_arrear':
            cursor.execute("USE 3_arrear_data")
            quary='delete from 3_arrear where year=%s and exam=%s and sem=%s'
            values=(year,exam,sem)
            cursor.execute(quary,values)
        elif arrear == 'two_arrear':
            cursor.execute("USE 2_arrear_data")
            quary='delete from 2_arrear where year=%s and exam=%s and sem=%s'
            values=(year,exam,sem)
            cursor.execute(quary,values) 
        elif arrear == 'one_arrear':
            cursor.execute("USE 1_arrear_data")
            quary='delete from 1_arrear where year=%s and exam=%s and sem=%s'
            values=(year,exam,sem)
            cursor.execute(quary,values)
        elif arrear == 'nil_arrear':
            cursor.execute("USE nil_arrear_data")
            quary='delete from nil_arrear where year=%s and exam=%s and sem=%s'
            values=(year,exam,sem)
            cursor.execute(quary,values)
        elif arrear == 'five_and_above':
            cursor.execute("USE 5_arrear_data")
            quary='delete from 5_arrear where year=%s and exam=%s'
            values=(year,exam)
            cursor.execute(quary,values)
        else:
            print("Invalid arrear type")

    finally:
        cnx.commit()
        cursor.close()
        cnx.close()
    return None
def clear_rec_data_overall():
    db_user = os.getenv("DB_USER")
    db_password = os.getenv("DB_PASSWORD")
    db_host = os.getenv("DB_HOST")
    cnx = pymysql.connect(
    cursorclass=pymysql.cursors.DictCursor,
    host=db_host,
    password=db_password,
    port=15274,
    user=db_user,)
    cursor = cnx.cursor()
    try:
        cursor.execute("USE overall_rec")
        query='delete from overall_data'
        cursor.execute(query)
        cnx.commit()
    finally:
        cursor.close()
        cnx.close()
    return None
def clear_rec_data():
    db_user = os.getenv("DB_USER")
    db_password = os.getenv("DB_PASSWORD")
    db_host = os.getenv("DB_HOST")
    cnx = pymysql.connect(
    cursorclass=pymysql.cursors.DictCursor,
    host=db_host,
    password=db_password,
    port=15274,
    user=db_user,)
    cursor = cnx.cursor()
    try:
        cursor.execute("USE status_rec")
        query='delete from status_data'
        cursor.execute(query)
        cnx.commit()
    finally:
        cursor.close()
        cnx.close()
    return None
def staff_del_data():
    db_user = os.getenv("DB_USER")
    db_password = os.getenv("DB_PASSWORD")
    db_host = os.getenv("DB_HOST")
    cnx = pymysql.connect(
    cursorclass=pymysql.cursors.DictCursor,
    host=db_host,
    password=db_password,
    port=15274,
    user=db_user,)
    cursor=cnx.cursor()
    query="USE all_data"
    cursor.execute(query)
    query1="DELETE FROM all_data1"
    cursor.execute(query1)
    cnx.commit()
    cursor.close()
    cnx.close()
def staff_del_data_overall():
    db_user = os.getenv("DB_USER")
    db_password = os.getenv("DB_PASSWORD")
    db_host = os.getenv("DB_HOST")
    cnx = pymysql.connect(
    cursorclass=pymysql.cursors.DictCursor,
    host=db_host,
    password=db_password,
    port=15274,
    user=db_user,)
    cursor=cnx.cursor()
    query="USE all_data_overall"
    cursor.execute(query)
    query1="DELETE FROM all_data_o"
    cursor.execute(query1)
    cnx.commit()
    cursor.close()
    cnx.close()
def message_del_data_overall():
    db_user = os.getenv("DB_USER")
    db_password = os.getenv("DB_PASSWORD")
    db_host = os.getenv("DB_HOST")
    cnx = pymysql.connect(
    cursorclass=pymysql.cursors.DictCursor,
    host=db_host,
    password=db_password,
    port=15274,
    user=db_user,)
    cursor=cnx.cursor()
    query="USE overall"
    cursor.execute(query)
    query1="DELETE FROM overall_data"
    cursor.execute(query1)
    cnx.commit()
    cursor.close()
    cnx.close()
def message_del_data():
    db_user = os.getenv("DB_USER")
    db_password = os.getenv("DB_PASSWORD")
    db_host = os.getenv("DB_HOST")
    cnx = pymysql.connect(
    cursorclass=pymysql.cursors.DictCursor,
    host=db_host,
    password=db_password,
    port=15274,
    user=db_user,)
    cursor=cnx.cursor()
    query="USE status"
    cursor.execute(query)
    query1="DELETE FROM status_data"
    cursor.execute(query1)
    cnx.commit()
    cursor.close()
    cnx.close()
def process_message_data():
    db_user = os.getenv("DB_USER")
    db_password = os.getenv("DB_PASSWORD")
    db_host = os.getenv("DB_HOST")
    cnx = pymysql.connect(
    cursorclass=pymysql.cursors.DictCursor,
    host=db_host,
    password=db_password,
    port=15274,
    user=db_user,)
    cursor = cnx.cursor()
    data1 = None 
    query="USE all_data"
    cursor.execute(query)
    query1="SELECT * FROM all_data1"
    cursor.execute(query1)
    data1 = cursor.fetchall()
    cursor.close()
    cnx.close()
    return data1
def process_message_data1():
    db_user = os.getenv("DB_USER")
    db_password = os.getenv("DB_PASSWORD")
    db_host = os.getenv("DB_HOST")
    cnx = pymysql.connect(
    cursorclass=pymysql.cursors.DictCursor,
    host=db_host,
    password=db_password,
    port=15274,
    user=db_user,)
    cursor = cnx.cursor()
    data2 = None 
    query="USE status"
    cursor.execute(query)
    query1="SELECT * FROM status_data"
    cursor.execute(query1)
    data2 = cursor.fetchall()
    cursor.close()
    cnx.close()
    return data2
def process_message_data2():
    db_user = os.getenv("DB_USER")
    db_password = os.getenv("DB_PASSWORD")
    db_host = os.getenv("DB_HOST")
    cnx = pymysql.connect(
    cursorclass=pymysql.cursors.DictCursor,
    host=db_host,
    password=db_password,
    port=15274,
    user=db_user,)
    cursor = cnx.cursor()
    data3 = None 
    query="USE status_rec"
    cursor.execute(query)
    query1="SELECT * FROM status_data"
    cursor.execute(query1)
    data3 = cursor.fetchall()
    cursor.close()
    cnx.close()
    return data3
def process_message_data_overall():
    db_user = os.getenv("DB_USER")
    db_password = os.getenv("DB_PASSWORD")
    db_host = os.getenv("DB_HOST")
    cnx = pymysql.connect(
    cursorclass=pymysql.cursors.DictCursor,
    host=db_host,
    password=db_password,
    port=15274,
    user=db_user,)
    cursor = cnx.cursor()
    data1 = None 
    query="USE all_data_overall"
    cursor.execute(query)
    query1="SELECT * FROM all_data_o"
    cursor.execute(query1)
    data1 = cursor.fetchall()
    cursor.close()
    cnx.close()
    return data1
def process_message_data_overall1():
    db_user = os.getenv("DB_USER")
    db_password = os.getenv("DB_PASSWORD")
    db_host = os.getenv("DB_HOST")
    cnx = pymysql.connect(
    cursorclass=pymysql.cursors.DictCursor,
    host=db_host,
    password=db_password,
    port=15274,
    user=db_user,)
    cursor = cnx.cursor()
    data2 = None 
    query="USE overall"
    cursor.execute(query)
    query1="SELECT * FROM overall_data"
    cursor.execute(query1)
    data2 = cursor.fetchall()
    cursor.close()
    cnx.close()
    return data2
def process_message_data_overall2():
    db_user = os.getenv("DB_USER")
    db_password = os.getenv("DB_PASSWORD")
    db_host = os.getenv("DB_HOST")
    cnx = pymysql.connect(
    cursorclass=pymysql.cursors.DictCursor,
    host=db_host,
    password=db_password,
    port=15274,
    user=db_user,)
    cursor = cnx.cursor()
    data3 = None 
    query="USE overall_rec"
    cursor.execute(query)
    query1="SELECT * FROM overall_data"
    cursor.execute(query1)
    data3 = cursor.fetchall()
    cursor.close()
    cnx.close()
    return data3
async def over_main(file_path, exam, year, cnx, cursor):
    print("Process started")
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    tasks = []
    for row in range(2, ws.max_row + 1):  # Skip the header row
        s_no = ws.cell(row=row, column=1).value
        name = ws.cell(row=row, column=2).value
        arrear_count = ws.cell(row=row, column=3).value
        phone_number = str(ws.cell(row=row, column=4).value)
        if arrear_count >= 5:
            message = f"Dear {name}, you have {arrear_count} Arrears in {exam.upper()}. Please take necessary action."
            query = "USE all_data_overall"
            cursor.execute(query)
            query1 = "INSERT INTO all_data_o (name, arrear_count, exam, year) VALUES (%s, %s, %s, %s)"
            values = (name, arrear_count, exam, year)
            cursor.execute(query1, values)
            cnx.commit()
            query2 = "USE 5_arrear_data"
            cursor.execute(query2)
            query3 = "INSERT INTO 5_arrear (name, arrear_count, exam, year) VALUES (%s, %s, %s, %s)"
            values = (name, arrear_count, exam, year)
            cursor.execute(query3, values)
            cnx.commit()
            phone_number = "+91" + phone_number
            tasks.append(send_sms_message1(name, arrear_count, exam, year, phone_number, message, cursor, cnx))
    await asyncio.gather(*tasks)
    print("Process completed")
async def main(file_path, exam, year, sem, cnx, cursor):
    print("Process started")
    cols = columns_read()
    data = read_excel_to_array(file_path)
    header = header_read(file_path)
    tasks = []
    output_file = os.path.join(os.getcwd(), 'templates','newsheet.xlsx')
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active
    ws.delete_cols(1, ws.max_column)
    ws.delete_rows(1, ws.max_row)
    ws.append(list(header)) 
    max_column = ws.max_column + 1
    ws.cell(row=1, column=max_column).value = "Arrear count"
    for i in range(0, len(data)):
        ws.append(data[i])
        count = 0
        subject = []  
        for j in range(3, cols-1):
            if int(data[i][j]) < 25:
                subject.append(header[j] + ' - ' + str(data[i][j])+'   (FAIL)')
                count += 1
            else:
                subject.append(header[j] + ' - ' + str(data[i][j])+'   (PASS)')
                
        
        ws.cell(row=i+2, column=max_column).value = count
        student_data = {
            "name": data[i][2],
            "phone_number": str(data[i][cols-1]),
            "subjects": subject,
            "arrear_count": count
        }
        qurey="USE all_data"
        cursor.execute(qurey)
        query1= "INSERT INTO all_data1 (name,arrear_count,sem,exam,year) VALUES (%s,%s, %s, %s, %s)"
        values = (data[i][2],count,sem,exam,year)
        cursor.execute(query1,values)
        cnx.commit()
        if count >= 3:
            name=data[i][2]
            phone_number = "+91" + student_data['phone_number']
            message = f"Dear {student_data['name']}, you have {count} Arrears in {exam.upper()}. Please take necessary action."
            for subject_detail in subject:
                message += f"\n{subject_detail}"
            tasks.append(send_sms_message(name,count,sem,exam,year,phone_number, message, cursor, cnx))
            qurey="USE 3_arrear_data"
            cursor.execute(qurey)
            query1= "INSERT INTO 3_arrear (name,arrear_count,sem,exam,year) VALUES (%s,%s, %s, %s, %s)"
            values = (data[i][2],count,sem,exam,year)
            cursor.execute(query1,values)
            cnx.commit()
        elif count == 2:
            qurey="USE 2_arrear_data"
            cursor.execute(qurey)
            query1= "INSERT INTO 2_arrear (name,arrear_count,sem,exam,year) VALUES (%s,%s, %s, %s, %s)"
            values = (data[i][2],count,sem,exam,year)
            cursor.execute(query1,values)
            cnx.commit()
        elif count == 1:
            qurey="USE 1_arrear_data"
            cursor.execute(qurey)
            query1= "INSERT INTO 1_arrear (name,arrear_count,sem,exam,year) VALUES (%s,%s, %s, %s, %s)"
            values = (data[i][2],count,sem,exam,year)
            cursor.execute(query1,values)
            cnx.commit()
        else:
            qurey="USE nil_arrear_data"
            cursor.execute(qurey)
            query1= "INSERT INTO nil_arrear (name,arrear_count,sem,exam,year) VALUES (%s,%s, %s, %s, %s)"
            values = (data[i][2],count,sem,exam,year)
            cursor.execute(query1,values)
            cnx.commit()
    wb.save(output_file)
    after_process()
    await asyncio.gather(*tasks)
    print("Process completed")
async def ESE_main(file_path, exam, year, sem, cnx, cursor):
    print("Process started")
    cols = columns_read()
    data = read_excel_to_array(file_path)
    header = header_read(file_path)
    tasks = []
    output_file = os.path.join(os.getcwd(), 'templates', 'newsheet.xlsx')
    
    if not os.path.exists(output_file):
        wb = openpyxl.Workbook()
    else:
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        for merged_range in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(merged_range))
        ws.delete_cols(1, ws.max_column)
        ws.delete_rows(1, ws.max_row)
    ws = wb.active
    ws.append(list(header)) 
    max_column = ws.max_column + 1
    ws.cell(row=1, column=max_column).value = "Arrear count"
    for i in range(1, len(data)):
        ws.append(data[i])
        count = 0
        subject = []  
        for j in range(3, cols-1):
            if isinstance(data[i][j],int):
                if data[i][j+1]=="RA" or data[i][j+1]=="ra" or data[i][j+1]=="A" or data[i][j+1]=="a":
                    subject.append(header[j] + ' - ' + str(data[i][j])+'   (FAIL)')
                    count+=1
                else:
                    subject.append(header[j] + ' - ' + str(data[i][j])+'   (PASS)')
                    continue
        ws.cell(row=i+2, column=max_column).value = count
        student_data = {
            "name": data[i][2],
            "phone_number": str(math.floor(data[i][cols-1])),
            "subjects": subject,
            "arrear_count": count
        }
        qurey="USE all_data"
        cursor.execute(qurey)
        query1= "INSERT INTO all_data1 (name,arrear_count,sem,exam,year) VALUES (%s,%s, %s, %s, %s)"
        values = (data[i][2],count,sem,exam,year)
        cursor.execute(query1,values)
        cnx.commit()
        if count >= 3:
            name=data[i][2]
            phone_number = "+91" + student_data['phone_number']
            message = f"Dear {student_data['name']}, you have {count} Arrears in {exam.upper()} End-semester Exam. Please take necessary action."
            for subject_detail in subject:
                message += f"\n{subject_detail}"
            tasks.append(send_sms_message(name,count,sem,exam,year,phone_number, message, cursor, cnx))
            qurey="USE 3_arrear_data"
            cursor.execute(qurey)
            query1= "INSERT INTO 3_arrear (name,arrear_count,sem,exam,year) VALUES (%s,%s, %s, %s, %s)"
            values = (data[i][2],count,sem,exam,year)
            cursor.execute(query1,values)
            cnx.commit()
        elif count == 2:
            qurey="USE 2_arrear_data"
            cursor.execute(qurey)
            query1= "INSERT INTO 2_arrear (name,arrear_count,sem,exam,year) VALUES (%s,%s, %s, %s, %s)"
            values = (data[i][2],count,sem,exam,year)
            cursor.execute(query1,values)
            cnx.commit()
        elif count == 1:
            qurey="USE 1_arrear_data"
            cursor.execute(qurey)
            query1= "INSERT INTO 1_arrear (name,arrear_count,sem,exam,year) VALUES (%s,%s, %s, %s, %s)"
            values = (data[i][2],count,sem,exam,year)
            cursor.execute(query1,values)
            cnx.commit()
        else:
            qurey="USE nil_arrear_data"
            cursor.execute(qurey)
            query1= "INSERT INTO nil_arrear (name,arrear_count,sem,exam,year) VALUES (%s,%s, %s, %s, %s)"
            values = (data[i][2],count,sem,exam,year)
            cursor.execute(query1,values)
            cnx.commit()
    wb.save(output_file)
    after_process_ese(file_path)
    await asyncio.gather(*tasks)
    print("Process completed")
app = Flask(__name__)
def get_or_create_eventloop():
    try:
        return asyncio.get_event_loop()
    except RuntimeError as ex:
        if "There is no current event loop in thread" in str(ex):
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            return asyncio.get_event_loop()

@app.route('/')
def index():
    return render_template('home.html')
@app.route('/home',methods=['POST'])
def home_route():
    return render_template('home.html')
@app.route('/about',methods=['POST'])
def about_route():
    return render_template('About.html')
@app.route('/login',methods=['POST'])
def login_route():
    return render_template('login.html')
@app.route('/back_overall',methods=['POST'])
def back_button_overall():
    flag=0
    try:
        db_user = os.getenv("DB_USER")
        db_password = os.getenv("DB_PASSWORD")
        db_host = os.getenv("DB_HOST")
        cnx = pymysql.connect(
            cursorclass=pymysql.cursors.DictCursor,
            host=db_host,
            password=db_password,
            port=15274,
            user=db_user,)
        cursor=cnx.cursor()
    except pymysql.MySQLError as e:
        flag=1
    if(flag==0):
        staff_del_data_overall()
        message_del_data_overall()
        return render_template('Staff.html',flag=flag)
    else:
        phone_no1=os.getenv("PH_NO1")
        ph_no1="+91"+str(phone_no1)
        phone_no2=os.getenv("PH_NO2")
        ph_no2="+91"+str(phone_no2)
        message="SERVER UNDER MAINTANCE"
        sendadmin1_msg(message,ph_no1)
        sendadmin2_msg(message,ph_no2)
        return render_template('Staff.html',flag=flag) 
@app.route('/back',methods=['POST'])
def back_button():
    flag=0
    try:
        db_user = os.getenv("DB_USER")
        db_password = os.getenv("DB_PASSWORD")
        db_host = os.getenv("DB_HOST")
        cnx = pymysql.connect(
            cursorclass=pymysql.cursors.DictCursor,
            host=db_host,
            password=db_password,
            port=15274,
            user=db_user,)
        cursor=cnx.cursor()
    except pymysql.MySQLError as e:
        flag=1
    if(flag==0):
        staff_del_data()
        message_del_data()
        return render_template('Staff.html',flag=flag)
    else:
        phone_no1=os.getenv("PH_NO1")
        ph_no1="+91"+str(phone_no1)
        phone_no2=os.getenv("PH_NO2")
        ph_no2="+91"+str(phone_no2)
        message="SERVER UNDER MAINTANCE"
        sendadmin1_msg(message,ph_no1)
        sendadmin2_msg(message,ph_no2)
        return render_template('Staff.html',flag=flag) 
@app.route('/back_hod',methods=['POST'])
def back_hod_button():
    return render_template('hod.html')
@app.route('/logout',methods=['POST'])
def logout_button():
    return render_template("login.html")
@app.route('/logout_data_overall',methods=['POST'])
def logout_data_overall():
    flag=0
    try:
        db_user = os.getenv("DB_USER")
        db_password = os.getenv("DB_PASSWORD")
        db_host = os.getenv("DB_HOST")
        cnx = pymysql.connect(
            cursorclass=pymysql.cursors.DictCursor,
            host=db_host,
            password=db_password,
            port=15274,
            user=db_user,)
        cursor=cnx.cursor()
    except pymysql.MySQLError as e:
        flag=1
    if(flag==0):
        staff_del_data_overall()
        message_del_data_overall()
        return render_template("login.html",flag=flag)
    else:
        phone_no1=os.getenv("PH_NO1")
        ph_no1="+91"+str(phone_no1)
        phone_no2=os.getenv("PH_NO2")
        ph_no2="+91"+str(phone_no2)
        message="SERVER UNDER MAINTANCE"
        sendadmin1_msg(message,ph_no1)
        sendadmin2_msg(message,ph_no2)
        return render_template('Staff.html',flag=flag)
@app.route('/logout_data',methods=['POST'])
def logout_data():
    flag=0
    try:
        db_user = os.getenv("DB_USER")
        db_password = os.getenv("DB_PASSWORD")
        db_host = os.getenv("DB_HOST")
        cnx = pymysql.connect(
            cursorclass=pymysql.cursors.DictCursor,
            host=db_host,
            password=db_password,
            port=15274,
            user=db_user,)
        cursor=cnx.cursor()
    except pymysql.MySQLError as e:
        flag=1
    if(flag==0):
        staff_del_data()
        message_del_data()
        return render_template("login.html",flag=flag)
    else:
        phone_no1=os.getenv("PH_NO1")
        ph_no1="+91"+str(phone_no1)
        phone_no2=os.getenv("PH_NO2")
        ph_no2="+91"+str(phone_no2)
        message="SERVER UNDER MAINTANCE"
        sendadmin1_msg(message,ph_no1)
        sendadmin2_msg(message,ph_no2)
        return render_template('Staff.html',flag=flag)
@app.route('/download')
def download_file():
    try:
        return send_file(os.path.join(os.getcwd(), 'templates', 'newsheet.xlsx'), as_attachment=True)
    except Exception as e:
        return str(e)
@app.route('/download_format')
def download_file_format():
    filename = request.args.get('filename', 'cae_format.xlsx')
    value = request.args.get('value', '')
    print(value)
    try:
        if(value=="cae1"):
            return send_file(os.path.join(os.getcwd(), 'templates', 'cae_format.xlsx'), as_attachment=True)
        elif(value=="cae2"):
            return send_file(os.path.join(os.getcwd(), 'templates', 'cae_format.xlsx'), as_attachment=True)
        elif(value=="ese"):
            return send_file(os.path.join(os.getcwd(), 'templates', 'ese_format.xlsx'), as_attachment=True)
        else:
            return send_file(os.path.join(os.getcwd(), 'templates', 'overall_format.xlsx'), as_attachment=True)
    except Exception as e:
        return str(e)
@app.route('/clear_rec_overall',methods=['POST'])
def clear_rec_overall():
    flag=0
    try:
        db_user = os.getenv("DB_USER")
        db_password = os.getenv("DB_PASSWORD")
        db_host = os.getenv("DB_HOST")
        cnx = pymysql.connect(
            cursorclass=pymysql.cursors.DictCursor,
            host=db_host,
            password=db_password,
            port=15274,
            user=db_user,)
        cursor=cnx.cursor()
    except pymysql.MySQLError as e:
        flag=1
    if(flag==0):
        clear_rec_data_overall()
        staff_del_data_overall()
        message_del_data_overall()
        return render_template('Staff.html',flag=flag)
    else:
        phone_no1=os.getenv("PH_NO1")
        ph_no1="+91"+str(phone_no1)
        phone_no2=os.getenv("PH_NO2")
        ph_no2="+91"+str(phone_no2)
        message="SERVER UNDER MAINTANCE"
        sendadmin1_msg(message,ph_no1)
        sendadmin2_msg(message,ph_no2)
        return render_template('Staff.html',flag=flag)
@app.route('/clear_rec',methods=['POST'])
def clear_rec():
    flag=0
    try:
        db_user = os.getenv("DB_USER")
        db_password = os.getenv("DB_PASSWORD")
        db_host = os.getenv("DB_HOST")
        cnx = pymysql.connect(
            cursorclass=pymysql.cursors.DictCursor,
            host=db_host,
            password=db_password,
            port=15274,
            user=db_user,)
        cursor=cnx.cursor()
    except pymysql.MySQLError as e:
        flag=1
    if(flag==0):
        clear_rec_data()
        staff_del_data()
        message_del_data()
        return render_template('Staff.html',flag=flag)
    else:
        phone_no1=os.getenv("PH_NO1")
        ph_no1="+91"+str(phone_no1)
        phone_no2=os.getenv("PH_NO2")
        ph_no2="+91"+str(phone_no2)
        message="SERVER UNDER MAINTANCE"
        sendadmin1_msg(message,ph_no1)
        sendadmin2_msg(message,ph_no2)
        return render_template('Staff.html',flag=flag)
@app.route('/clear_data',methods=['POST'])
def clear():
    arrear=request.form['arrear']
    year=request.form['year']
    exam=request.form['exam']
    sem=request.form['sem']
    clear_data(arrear,year,exam,sem)
    return render_template('hod.html')
@app.route('/login_page', methods=['POST'])
def login_page():
    login_user = request.form['login_user']
    email = request.form['email_user']
    password = request.form['password_user']
    loop = get_or_create_eventloop()
    stat = loop.run_until_complete(login_main(login_user, email,password))
    if stat == True:
        return render_template('hod.html')
    elif stat == False:
        return render_template('Staff.html')
    else:
        return render_template('login.html')
@app.route('/hod_page',methods=['POST'])
def hod_data():
    flag=0
    try:
        db_user = os.getenv("DB_USER")
        db_password = os.getenv("DB_PASSWORD")
        db_host = os.getenv("DB_HOST")
        cnx = pymysql.connect(
            cursorclass=pymysql.cursors.DictCursor,
            host=db_host,
            password=db_password,
            port=15274,
            user=db_user,)
        cursor=cnx.cursor()
    except pymysql.MySQLError as e:
        flag=1
    if(flag==0):
        exam = request.form['form_sheet']
        year = request.form['year']  # Get year from form input
        sem = request.form['sem']  # Get semester from form input
        arrear=request.form['arrears']
        if(exam=="overall"):
            data=process_hod_data_overall(year,exam,arrear,cnx,cursor)
            cursor.close()
            cnx.close()
            return render_template('data_overall.html',data=data,arrear=arrear,exam=exam,year=year)
        else:
            data=process_hod_data(year, sem, exam, arrear, cnx, cursor)
            cursor.close()
            cnx.close()
            return render_template('data.html',data=data,arrear=arrear,exam=exam,year=year,sem=sem)
    else:
        phone_no1=os.getenv("PH_NO1")
        ph_no1="+91"+str(phone_no1)
        phone_no2=os.getenv("PH_NO2")
        ph_no2="+91"+str(phone_no2)
        message="SERVER UNDER MAINTANCE"
        sendadmin1_msg(message,ph_no1)
        sendadmin2_msg(message,ph_no2)
        return render_template('Staff.html',flag=flag)
@app.route('/upload', methods=['POST'])
def upload_marks():
    flag=0
    try:
        db_user = os.getenv("DB_USER")
        db_password = os.getenv("DB_PASSWORD")
        db_host = os.getenv("DB_HOST")
        cnx = pymysql.connect(
            cursorclass=pymysql.cursors.DictCursor,
            host=db_host,
            password=db_password,
            port=15274,
            user=db_user,)
        cursor=cnx.cursor()
    except pymysql.MySQLError as e:
        flag=1
    if request.method == 'POST':
        exam = request.form['form_sheet']
        year = request.form['year']  # Get year from form input
        sem = request.form['sem']  # Get semester from form input
        file = request.files['file']
        file.save(os.path.join(os.getcwd(), 'Marks1.xlsx'))
        if(flag==0):
            if exam=="cae1" or exam=="cae2":
                loop = get_or_create_eventloop()
                loop.run_until_complete(main('Marks1.xlsx', exam, year, sem, cnx, cursor))
                cursor.close()
                cnx.close()
                data1=process_message_data()
                data2=process_message_data1()
                data3=process_message_data2()
                return render_template('message.html',data1=data1,data2=data2,data3=data3)
            elif exam=="ese":
                loop=get_or_create_eventloop()
                loop.run_until_complete(ESE_main('Marks1.xlsx',exam,year,sem, cnx, cursor))
                cursor.close()
                cnx.close()
                data1=process_message_data()
                data2=process_message_data1()
                data3=process_message_data2()
                return render_template('message.html',data1=data1,data2=data2,data3=data3)
            else:
                loop = get_or_create_eventloop()
                loop.run_until_complete(over_main('Marks1.xlsx', exam, year, cnx, cursor))
                cursor.close()
                cnx.close()
                data1=process_message_data_overall()
                data2=process_message_data_overall1()
                data3=process_message_data_overall2()
                print(data1,data2,data3)
                return render_template('message_overall.html',data1=data1,data2=data2,data3=data3)
        else:
            phone_no1=os.getenv("PH_NO1")
            ph_no1="+91"+str(phone_no1)
            phone_no2=os.getenv("PH_NO2")
            ph_no2="+91"+str(phone_no2)
            message="SERVER UNDER MAINTANCE"
            sendadmin1_msg(message,ph_no1)
            sendadmin2_msg(message,ph_no2)
            return render_template('Staff.html',flag=flag)
# Run the Flask application
if __name__ == '__main__':
    app.run(debug=True)
